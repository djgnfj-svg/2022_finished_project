import sys
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import QDate
from PyQt5 import uic
from openpyxl import Workbook, load_workbook
from functools import partial

from Utile      import *
from Time_Utiles import *
Review_UI = resource_path('UI/Review_UI.ui')

class Review(QDialog):
    def __init__(self):
        QDialog.__init__(self,None)
        uic.loadUi(Review_UI,self)
        self.page = 0
        self.pages = []
        #기본 UI
        load_wb, load_ws = load_Update()
        self.page_max, self.pages = get_page_max(load_ws)
        self.Descrip.setPlainText(load_ws.cell(self.pages[self.page],2).value)
        self.Date.setPlainText(load_ws.cell(self.pages[self.page],3).value)

        self.BackButton.setDisabled(True)
        if (self.page_max == 1):
            self.NextButton.setDisabled(True)
        #버튼 액션
        self.CheckButton.clicked.connect(partial(self.pushCheck, load_ws))
        self.CancelButton.clicked.connect(partial(self.pushCancel,load_ws,load_wb))
        self.BackButton.clicked.connect(partial(self.pushBack,load_ws))
        self.NextButton.clicked.connect(partial(self.pushNext, load_ws))

    def pushCheck(self, load_ws):
        Title = load_ws.cell(self.pages[self.page],1).value
        self.Title.setPlainText(Title)

    def pushBack(self, load_ws):
        self.page -= 1
        self.Descrip.setPlainText(load_ws.cell(self.pages[self.page],2).value)
        self.Date.setPlainText(load_ws.cell(self.pages[self.page],3).value)

        self.Title.clear()
        if (self.page >= len(self.pages)-1):
            self.NextButton.setDisabled(True)
        else:
            self.NextButton.setEnabled(True)
        if (self.page < 1):
            self.BackButton.setDisabled(True)
        else:
            self.BackButton.setEnabled(True)

    def pushNext(self, load_ws):
        self.page += 1
        self.Descrip.setPlainText(load_ws.cell(self.pages[self.page],2).value)
        self.Date.setPlainText(load_ws.cell(self.pages[self.page],3).value)
        self.Title.clear()

        if (self.page >= len(self.pages)-1):
            self.NextButton.setDisabled(True)
        else:
            self.NextButton.setEnabled(True)

        if (self.page < 1):
            self.BackButton.setDisabled(True)
        else:
            self.BackButton.setEnabled(True)

    def pushCancel(self,load_ws,load_wb):
        self.mes = QMessageBox()
        answer = self.mes.question(self, '복습완료?', \
        '복습을 완료하셨습니까?',QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if (answer == QMessageBox.Yes):
            for i in self.pages:
                load_ws.cell(i,4).value += 1
            load_wb.save(resource_path("./excel/test_open.xlsx"))
        else :
            self.mes.close()
        self.close()

