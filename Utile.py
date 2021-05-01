import sys
import os
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import QDate
from PyQt5 import uic
from openpyxl import Workbook, load_workbook
from functools import partial

import sys
import os
 
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def load_Update():
    load_wb = load_workbook(resource_path('./excel/test_open.xlsx'))
    load_ws = load_wb['Sheet']
    return load_wb, load_ws

# 엑셀의 유효한 값까지 일어주는 함수
def get_line_Line_number(load_ws):
    for i in range(1,load_ws.max_row+1):
        if (load_ws.cell(i, 1).value):
            pass
        else :
            break
    return i

def check_duplicates(Title, load_ws, Line_number):
    duplicates = False
    for i in range(1, Line_number+1):
        # 이면서 임시 타이틀과 같아야함 임시타ㅇ
        if Title in str(load_ws.cell(i, 1).value):
            duplicates = True
            break
    return duplicates

def Write_excel(add_list, load_ws, Line_number):
    for i, add in enumerate(add_list) : # 5개 제목 내용 시간 스택 복습한거 테스트용
        load_ws.cell(Line_number, i+1, add)

