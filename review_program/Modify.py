import sys
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import QDate
from PyQt5 import uic
from openpyxl import Workbook, load_workbook
from functools import partial
from datetime import datetime
from Utile import *

import re
import sys
import io


Modify_UI = resource_path('UI/Modify_UI.ui')
Write_UI = resource_path('UI/Write_UI.ui')

class Modify(QDialog):
    def __init__(self):
        super().__init__()
        uic.loadUi(Modify_UI,self)

        self.load_wb, self.load_ws = load_Update()
        self.Search_list = []
        self.all_rows = get_line_Line_number(self.load_ws)

        self.list_init()
        # 여까지

        self.view.itemDoubleClicked.connect(self.DoubleClicked)
        self.Search_btn.clicked.connect(partial(self.push_Search, self.Search_list))
        self.Search_text.editingFinished.connect(partial(self.push_Search, self.Search_list))

        self.Cancel_btn.clicked.connect(self.push_Cancel)

    def push_Search(self, Search_list):
        Title = self.Search_text.text()
        if Title == '':
            self.list_init()
        else :
            self.view.clear()
            for i, search_A in enumerate(self.Search_list,start=1) :
                search_A.rstrip()
                if (search_A.find(Title) >= 0):
                    item = QListWidgetItem()
                    item.setText(str(self.load_ws.cell(i, 1).value) +'\t' + str(self.load_ws.cell(i, 2).value)[:10])
                    self.view.addItem(item)

    def list_init(self):
        self.view.clear()
        for i in range(1, self.all_rows+1):
            if(self.load_ws.cell(i, 1).value):
                item = QListWidgetItem()
                item.setText(str(self.load_ws.cell(i, 1).value) +'\t' + str(self.load_ws.cell(i, 2).value)[:10])
                self.Search_list.append(self.load_ws.cell(i,1).value)
                self.view.addItem(item)
            else:
                break

    def DoubleClicked(self):
        temp = self.view.currentItem().text()
        Title = temp.split('\t')[0]

        dlg = Modify_write(Title)
        dlg.exec_()
        self.load_wb, self.load_ws = load_Update()
        self.list_init()
    
    def push_Cancel(self):
        self.close()

class Modify_write(QDialog):
    def __init__(self, Title):

        super().__init__()
        uic.loadUi(Write_UI,self)

        load_wb, load_ws = load_Update()
        ## 기존에 있던 내용 적기
        
        self.find_row = 0
        self.all_rows = get_line_Line_number(load_ws) 
        for i in range(1, self.all_rows): # 엑셀의 마지막 행까지 분석하는 숫자사용하자
            if(str(load_ws.cell(i, 1).value) == Title):
                self.find_row = i
                break
            else:
                continue
        self.Title.setPlainText(load_ws.cell(self.find_row,1).value)
        self.Descrip.setPlainText(load_ws.cell(self.find_row,2).value)
        self.Date.setPlainText(datetime.today().strftime("%D"))

        self.temp_title = Title = self.Title.toPlainText()

        self.SaveButton.clicked.connect(self.pushSave)
        self.CancelButton.clicked.connect(self.pushCancel)

    def pushSave(self):
        Title = self.Title.toPlainText()
        Descrip = self.Descrip.toPlainText()
        now = self.Date.toPlainText()
        review_stack = 1

        load_wb, load_ws = load_Update()

        self.Line_number = get_line_Line_number(load_ws) # 엑셀이 유효한 값의 마지막을 반환

        add_list = [Title, Descrip, now,review_stack]
        if ((check_duplicates(Title, load_ws, self.Line_number)) == True) and \
            Title != self.temp_title :
            self.mes = QMessageBox()
            Reply = self.mes.information(self, '제목 중복', \
                '이미 같은 제목이 있습니다.')
            return 0
        Write_excel(add_list, load_ws, self.Line_number)

        load_wb.save(resource_path("./excel/test_open.xlsx"))
        self.close()

    def pushCancel(self):
       self.close()